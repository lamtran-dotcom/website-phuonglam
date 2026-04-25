from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "file sản phẩm từ shopee"
INFO_FILE = SOURCE_DIR / "thong tin sản phẩm.xlsx"
IMAGE_FILE = SOURCE_DIR / "ảnh sản phẩm.xlsx"
DESCRIPTION_FILE = SOURCE_DIR / "mô tả sản phẩm.xlsx"
WEIGHT_FILE = SOURCE_DIR / "cân nặng.xlsx"
OUTPUT_FILE = SOURCE_DIR / "shopee-products-import.json"

SEO_OVERRIDES = {
    "57653126590": {
        "name": "Quế Thanh Yên Bái Xuất Khẩu - Quế Khô Nguyên Thanh Thơm Đậm",
        "shortDesc": "Quế thanh Yên Bái xuất khẩu, chọn từ quế già cây, thơm nồng ấm và nhiều tinh dầu. Phù hợp nấu phở, bò kho, món hầm, pha trà quế và tạo hương tự nhiên.",
        "description": "QUẾ THANH YÊN BÁI XUẤT KHẨU\n\n✔ Tổng quan sản phẩm:\nQuế thanh Yên Bái xuất khẩu là dòng quế khô nguyên thanh được chọn từ quế già cây, vỏ dày, chắc, mùi thơm nồng ấm và hàm lượng tinh dầu tự nhiên cao.\nSản phẩm phù hợp cho gia đình cần quế thanh sạch để nấu ăn, pha trà thảo mộc, làm đồ uống ấm hoặc dùng decor tạo hương tự nhiên.\n\n✔ Điểm nổi bật:\n- Quế Yên Bái thơm đậm, rõ mùi, bền hương.\n- Thanh quế dày, chắc, nhiều tinh dầu tự nhiên.\n- Hậu vị ngọt sâu, cay nhẹ, ít chát.\n- Đã cạo vỏ sạch, màu nâu đỏ đẹp, tiện sử dụng.\n- Dễ bảo quản, dùng được lâu khi buộc kín túi và để nơi khô thoáng.\n\n✔ Ứng dụng phổ biến:\n- Gia vị nấu ăn: dùng cho phở, bò kho, món hầm, món nướng, nước dùng và các món cần hương quế ấm.\n- Pha trà và đồ uống: pha trà quế mật ong, trà quế gừng, trà thảo mộc hoặc trang trí đồ uống nóng.\n- Tạo hương tự nhiên: đặt trong phòng, tủ quần áo, góc decor hoặc kết hợp cùng nến, thảo mộc khô.\n- Quà tặng thảo mộc: hình thức sạch, đẹp, hương thơm tự nhiên, phù hợp làm quà nhỏ tinh tế.\n\n✔ Thông tin sản phẩm:\n- Tên sản phẩm: Quế thanh Yên Bái xuất khẩu\n- Dạng sản phẩm: quế khô nguyên thanh, đã cạo vỏ sạch\n- Xuất xứ: Yên Bái, Việt Nam\n- Phân loại: 50g, 100g, 200g\n- Cân nặng vận chuyển: 200g\n- Bảo quản: nơi khô ráo, thoáng mát, tránh ẩm, buộc kín túi zip sau khi dùng.\n\n📍 Cam kết:\n- Sản phẩm được chọn lọc từ nguồn quế tự nhiên.\n- Hình ảnh và mô tả được chuẩn hóa rõ ràng để khách dễ chọn đúng phân loại.\n- Hỗ trợ tư vấn cách dùng quế thanh cho nấu ăn, pha trà và tạo hương.",
        "usage": "HƯỚNG DẪN SỬ DỤNG QUẾ THANH\n\n✔ Nấu ăn:\nCho 1-2 thanh quế vào phở, bò kho, món hầm hoặc nước dùng. Vớt ra khi hương quế đã đủ thơm để món ăn không bị gắt mùi.\n\n✔ Pha trà:\nHãm 1 thanh quế nhỏ với nước nóng 5-10 phút. Có thể thêm mật ong, gừng, táo đỏ hoặc thảo mộc tùy khẩu vị.\n\n✔ Tạo hương tự nhiên:\nĐặt vài thanh quế ở bàn trà, kệ decor, tủ quần áo hoặc túi thơm để tạo mùi ấm nhẹ.\n\n✔ Bảo quản:\nBuộc kín túi sau khi dùng, để nơi khô ráo, tránh nắng gắt và tránh ẩm để giữ mùi thơm lâu hơn.",
    },
    "57058955812": {
        "categoryId": "nen-thom",
        "name": "Set 50 Nến Tealight Trang Trí - Vỏ Trong/Vỏ Nhôm Cháy Ổn Định",
        "shortDesc": "Set 50 viên nến tealight nhỏ gọn, ánh sáng ấm, phù hợp trang trí bàn tiệc, phòng ngủ, spa, thiền thư giãn và dùng kèm đèn xông tinh dầu hoặc bếp xông thảo mộc.",
        "description": "SET 50 NẾN TEALIGHT TRANG TRÍ\n\n✔ Tổng quan sản phẩm:\nSet 50 viên nến tealight được thiết kế nhỏ gọn, dễ dùng và cháy ổn định, phù hợp để tạo ánh sáng ấm cho bàn tiệc, phòng ngủ, spa, góc thiền hoặc không gian thư giãn tại nhà.\nSản phẩm có 2 phân loại để khách chọn theo nhu cầu: vỏ trong suốt đẹp khi decor và vỏ nhôm phù hợp hơn khi dùng để xông tinh dầu, xông thảo mộc liên tục.\n\n✔ Điểm nổi bật:\n- Set 50 viên tiện lợi, dễ dự trữ và sử dụng hằng ngày.\n- Ánh sáng ấm, dịu mắt, tạo cảm giác thư giãn và chill vào buổi tối.\n- Kích thước nhỏ gọn, dễ đặt trong ly nến, đèn xông tinh dầu hoặc phụ kiện decor.\n- Vỏ trong suốt cho hiệu ứng ánh sáng đẹp, phù hợp chụp ảnh, trang trí bàn tiệc và decor phòng.\n- Vỏ nhôm chịu nhiệt tốt hơn, phù hợp dùng kèm bếp xông thảo mộc hoặc đèn xông tinh dầu trong thời gian dài.\n\n✔ Ứng dụng phổ biến:\n- Trang trí bàn tiệc, sinh nhật, đám cưới, quán cà phê, spa hoặc phòng ngủ.\n- Dùng cùng đèn xông tinh dầu để khuếch tán hương thơm nhẹ nhàng.\n- Dùng cùng bếp xông thảo mộc để tạo nhiệt ổn định khi xông nhà.\n- Setup thiền, yoga, đọc sách hoặc thư giãn buổi tối.\n\n✔ Thông tin sản phẩm:\n- Tên sản phẩm: Set 50 nến tealight trang trí\n- Phân loại: 50 viên vỏ trong suốt, 50 viên vỏ nhôm\n- Cân nặng vận chuyển: 2000g\n- Dạng sản phẩm: nến tealight viên nhỏ\n- Phù hợp: decor, spa, thiền, xông tinh dầu, xông thảo mộc\n\n✔ Lưu ý quan trọng:\n- Vỏ trong suốt đẹp hơn khi dùng để trang trí/decor.\n- Nếu dùng để xông tinh dầu hoặc xông thảo mộc liên tục, nên chọn loại vỏ nhôm để chịu nhiệt tốt hơn và bền hơn.\n- Luôn đặt nến trên bề mặt phẳng, chịu nhiệt và tránh xa vật dễ cháy.",
        "usage": "HƯỚNG DẪN SỬ DỤNG NẾN TEALIGHT\n\n✔ Dùng để trang trí:\nĐặt nến trong ly nến, khay decor hoặc khu vực bàn tiệc để tạo ánh sáng ấm và không gian thư giãn.\n\n✔ Dùng với đèn xông tinh dầu:\nĐặt nến vào khoang đốt của đèn xông, sau đó thêm nước và tinh dầu vào khay phía trên theo hướng dẫn của từng loại đèn.\n\n✔ Dùng với bếp xông thảo mộc:\nƯu tiên chọn phân loại vỏ nhôm nếu cần đốt liên tục để xông thảo mộc hoặc giữ nhiệt lâu hơn.\n\n✔ An toàn khi sử dụng:\nKhông để nến gần rèm, giấy, gỗ mỏng hoặc vật dễ cháy. Không di chuyển nến khi đang cháy và luôn tắt nến trước khi rời khỏi phòng.\n\n✔ Bảo quản:\nĐể nến nơi khô ráo, thoáng mát, tránh nắng nóng trực tiếp để giữ form nến đẹp và hạn chế chảy sáp.",
    },
    "56102764086": {
        "categoryId": "phu-kien",
        "name": "Dĩa Lót Bếp Xông Đất Nung - Chịu Nhiệt, Giữ Sạch Mặt Bàn",
        "shortDesc": "Dĩa lót đất nung tự nhiên dùng cho bếp xông thảo mộc, bồ kết và nến tealight. Giúp cách nhiệt, giữ vệ sinh mặt bàn và hoàn thiện bộ bếp xông mộc mạc.",
        "description": "DĨA LÓT BẾP XÔNG ĐẤT NUNG\n\n✔ Tổng quan sản phẩm:\nDĩa lót bếp xông đất nung là phụ kiện cần có khi sử dụng bếp xông thảo mộc, bồ kết hoặc đèn xông dùng nến.\nSản phẩm giúp bộ bếp xông hoàn thiện hơn, đồng thời hỗ trợ giữ vệ sinh và tăng độ an toàn khi đặt bếp trên mặt bàn.\n\n✔ Công dụng nổi bật:\n- Cách nhiệt an toàn, hạn chế nhiệt từ đáy bếp tiếp xúc trực tiếp với mặt bàn gỗ, kính hoặc đá.\n- Hứng tàn nến, sáp hoặc vụn thảo mộc rơi vãi để khu vực xông luôn gọn sạch.\n- Màu đất nung mộc mạc, ton-sur-ton với bếp xông, phù hợp không gian spa, phòng khách, bàn trà hoặc bàn thờ gia tiên.\n- Dễ kết hợp với bếp xông size 13cm và 16cm.\n\n✔ Thông tin sản phẩm:\n- Tên sản phẩm: Dĩa lót bếp xông đất nung\n- Chất liệu: đất nung tự nhiên, chịu nhiệt tốt\n- Màu sắc: nâu đỏ gốm mộc\n- Kích thước tương thích: bếp xông đường kính khoảng 13cm - 16cm\n- Xuất xứ: Việt Nam\n- Cân nặng vận chuyển: 200g\n\n✔ Lưu ý quan trọng:\n- Sản phẩm chỉ bao gồm dĩa lót, không bao gồm bếp xông, nến hoặc thảo mộc.\n- Do là gốm đất nung, màu sắc có thể chênh lệch đậm nhạt nhẹ tùy mẻ nung.\n- Nên đặt dĩa trên bề mặt phẳng, khô ráo trước khi đặt bếp xông lên trên.\n\n📍 Cam kết:\n- Hàng đúng mô tả, đóng gói chống sốc kỹ khi vận chuyển.\n- Hỗ trợ đổi trả nếu sản phẩm bị bể vỡ do quá trình vận chuyển.",
        "usage": "HƯỚNG DẪN SỬ DỤNG DĨA LÓT BẾP XÔNG\n\n✔ Cách dùng:\nĐặt dĩa lót trên mặt bàn phẳng, sau đó đặt bếp xông thảo mộc, bếp xông bồ kết hoặc phụ kiện xông lên giữa dĩa.\n\n✔ Khi dùng với nến:\nĐặt nến tealight vào đúng vị trí của bếp xông. Dĩa lót bên dưới giúp hứng tàn, sáp hoặc vụn thảo mộc nếu có rơi xuống.\n\n✔ Vệ sinh:\nSau khi bếp nguội hoàn toàn, lau dĩa bằng khăn khô hoặc khăn ẩm nhẹ. Không ngâm nước quá lâu để giữ độ bền của đất nung.\n\n✔ Bảo quản:\nĐể nơi khô thoáng, tránh va đập mạnh. Khi không dùng, có thể đặt chung với bộ bếp xông để giữ trọn bộ gọn gàng.",
    },
    "55553449576": {
        "categoryId": "combo",
        "name": "Bộ Bếp Xông Thảo Mộc Đất Nung Full Set - Kèm Nến, Dĩa Lót & Thảo Dược",
        "shortDesc": "Bộ bếp xông thảo mộc đất nung full set 5 món, có bếp niêu, đế nến, dĩa lót, 10 viên nến tealight và thảo mộc tự nhiên. Phù hợp xông nhà, khử mùi, thư giãn và tạo hương ấm cho không gian sống.",
        "description": "BỘ BẾP XÔNG THẢO MỘC ĐẤT NUNG FULL SET\n\n✔ Tổng quan sản phẩm:\nBộ bếp xông thảo mộc Phương Lâm là combo xông nhà tiện lợi, kết hợp bếp niêu đất nung mộc mạc với nến tealight và thảo mộc tự nhiên.\nSản phẩm phù hợp cho khách muốn mua một bộ đầy đủ để dùng ngay, không cần chọn lẻ từng phụ kiện.\n\n✔ Bộ sản phẩm bao gồm:\n- Bếp niêu đất nung size 13cm hoặc 16cm tùy phân loại.\n- Đế đựng nến đất nung giúp thay nến dễ hơn và an toàn hơn khi sử dụng.\n- Dĩa lót bếp giúp cách nhiệt, giữ vệ sinh mặt bàn và hoàn thiện bộ bếp.\n- 10 viên nến tealight cháy khoảng 4 giờ mỗi viên.\n- Gói thảo mộc xông tự nhiên được phối theo từng nhu cầu: tẩy uế, khử mùi, thư giãn, thanh lọc không khí, xua muỗi, giảm cảm hoặc xông nhà mới.\n\n✔ Công dụng nổi bật:\n- Xông nhà mới, xông phòng làm việc, phòng khách, phòng ngủ hoặc khu vực cần làm sạch mùi.\n- Khử mùi ẩm mốc, mùi thức ăn và tạo hương thảo mộc ấm dễ chịu.\n- Hỗ trợ không gian thư giãn, thiền, đọc sách, spa tại nhà hoặc nghỉ ngơi cuối ngày.\n- Thiết kế đất nung mộc mạc, dễ decor cùng bàn trà, kệ gỗ, góc thờ hoặc không gian phong cách tự nhiên.\n\n✔ Chọn phân loại theo nhu cầu:\n- Size 13cm: nhỏ gọn, phù hợp phòng nhỏ, bàn làm việc, phòng ngủ hoặc khách mới bắt đầu dùng bếp xông.\n- Size 16cm: lòng bếp rộng hơn, phù hợp phòng khách, không gian lớn hơn hoặc nhu cầu xông thường xuyên.\n- Bộ Tẩy Uế: phù hợp xông nhà mới, cuối năm, đầu năm, rằm hoặc mùng 1.\n- Bộ Khử Mùi: phù hợp khu vực bếp, phòng kín, phòng có mùi ẩm hoặc mùi thức ăn.\n- Bộ Thư Giãn: phù hợp phòng ngủ, spa tại nhà, thiền hoặc nghỉ ngơi buổi tối.\n- Bộ Thanh Lọc Không Khí: phù hợp dùng hằng ngày để tạo không gian thơm sạch tự nhiên.\n- Bộ Xua Muỗi: phù hợp khu vực có muỗi, côn trùng hoặc phòng cần mùi thảo mộc mạnh hơn.\n- Bộ Giảm Cảm: phù hợp khi cần hương thảo mộc ấm, dễ chịu cho mùa lạnh hoặc lúc cơ thể mệt mỏi.\n- Bộ Xông Nhà Mới: phù hợp khai trương, nhập trạch, dọn nhà hoặc làm mới không gian.\n\n✔ Thông tin sản phẩm:\n- Tên sản phẩm: Bộ bếp xông thảo mộc đất nung full set\n- Chất liệu bếp: đất nung chịu nhiệt\n- Phân loại kích thước: 13cm và 16cm\n- Số nến đi kèm: 10 viên tealight\n- Cân nặng vận chuyển: 2000g\n- Phù hợp: xông nhà, khử mùi, tẩy uế, thư giãn, decor, thanh lọc không khí\n\n✔ Lưu ý quan trọng:\n- Khi xông thảo mộc, dùng phương pháp nướng khô để thảo mộc tỏa hương, không cần cho nước.\n- Đặt bếp trên dĩa lót, bề mặt phẳng và cách xa vật dễ cháy.\n- Không chạm tay vào bếp khi đang đốt nến vì đất nung giữ nhiệt lâu.\n- Để xa tầm tay trẻ em và vật nuôi trong suốt quá trình sử dụng.",
        "usage": "HƯỚNG DẪN SỬ DỤNG BỘ BẾP XÔNG THẢO MỘC\n\n✔ Bước 1: Chuẩn bị bếp\nĐặt dĩa lót trên mặt bàn phẳng, sau đó đặt bếp niêu đất nung lên giữa dĩa.\n\n✔ Bước 2: Đốt nến\nĐặt 1 viên nến tealight vào đế đựng nến, châm lửa rồi đưa đế nến vào lòng bếp qua cửa lò.\n\n✔ Bước 3: Cho thảo mộc\nCho một lượng thảo mộc vừa đủ vào lòng niêu phía trên. Không cần thêm nước, để thảo mộc được làm nóng khô và tỏa hương tự nhiên.\n\n✔ Bước 4: Tận hưởng hương thơm\nĐặt bếp ở nơi thông thoáng nhẹ để hương thảo mộc lan đều. Có thể thêm thảo mộc nếu muốn mùi rõ hơn.\n\n✔ Sau khi dùng:\nChờ bếp nguội hoàn toàn rồi lấy phần thảo mộc đã xông ra. Lau bếp và dĩa lót bằng khăn khô hoặc khăn ẩm nhẹ.\n\n✔ An toàn:\nKhông di chuyển bếp khi nến đang cháy. Không để gần rèm, giấy, gỗ mỏng hoặc vật dễ bắt lửa. Luôn tắt nến trước khi rời khỏi phòng.",
        "variantNameMap": {
            "16cm,BỘ TẨY UẾ - FULL": "16cm - Bộ Tẩy Uế Full Set",
            "16cm,BỘ GIẢM CẢM - FULL": "16cm - Bộ Giảm Cảm Full Set",
            "13cm,BỘ XUA MUỖI - FULL": "13cm - Bộ Xua Muỗi Full Set",
            "16cm,BÔ XÔNG NHÀ MỚI-FULL": "16cm - Bộ Xông Nhà Mới Full Set",
            "16cm,BỘ XUA MUỖI - FULL": "16cm - Bộ Xua Muỗi Full Set",
            "13cm,BỘ KHỬ MÙI - FULL": "13cm - Bộ Khử Mùi Full Set",
            "13cm,BỘ THƯ GIÃN - FULL": "13cm - Bộ Thư Giãn Full Set",
            "13cm,BỘ THANH LỌC KK-FULL": "13cm - Bộ Thanh Lọc Không Khí Full Set",
            "13cm,BÔ XÔNG NHÀ MỚI-FULL": "13cm - Bộ Xông Nhà Mới Full Set",
            "16cm,BỘ KHỬ MÙI - FULL": "16cm - Bộ Khử Mùi Full Set",
            "16cm,BỘ THANH LỌC KK-FULL": "16cm - Bộ Thanh Lọc Không Khí Full Set",
            "13cm,BỘ GIẢM CẢM - FULL": "13cm - Bộ Giảm Cảm Full Set",
            "16cm,BỘ THƯ GIÃN - FULL": "16cm - Bộ Thư Giãn Full Set",
            "13cm,BỘ TẨY UẾ - FULL": "13cm - Bộ Tẩy Uế Full Set"
        },
    }
}


def clean_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def clean_multiline(value) -> str:
    text = str(value or "").replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"(?im)^\s*#\S+.*$", "", text)
    text = re.sub(r"(?im)^.*từ\s*khóa\s*seo.*$", "", text)
    return text.strip()


def to_number(value):
    if value in (None, ""):
        return None
    text = re.sub(r"[^\d.]", "", str(value))
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_grams(text: str):
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*(kg|g)\b", text.lower())
    if not match:
        return None
    value = float(match.group(1).replace(",", "."))
    return int(value * 1000) if match.group(2) == "kg" else int(value)


def infer_category(name: str, shopee_category: str = "") -> str:
    haystack = f"{name} {shopee_category}".lower()
    rules = [
        ("nu-tram", ["nụ trầm", "nhang nụ", "trầm"]),
        ("nen-ly", ["nến ly", "ly nến"]),
        ("nen-tru", ["nến trụ"]),
        ("combo", ["combo", "full set", "bộ"]),
        ("bep-xong", ["bếp", "đèn xông", "máy xông"]),
        ("phu-kien", ["phụ kiện", "dĩa", "đĩa", "nắp", "đế", "hột quẹt", "quẹt", "cốc kê", "ly đựng"]),
        ("thao-moc-xong", ["thảo mộc", "lá dứa", "quế", "sả", "gừng", "ngải cứu"]),
        ("tinh-dau", ["tinh dầu", "essential oils"]),
        ("nen-thom", ["nến tealight", "tealight", "candles", "nến"]),
    ]
    for category, keywords in rules:
        if any(keyword in haystack for keyword in keywords):
            return category
    return "phu-kien"


def read_info_rows():
    wb = load_workbook(INFO_FILE, read_only=True, data_only=True)
    ws = wb.active
    products = defaultdict(list)
    for row in ws.iter_rows(min_row=7, values_only=True):
        product_id = clean_text(row[0] if len(row) > 0 else "")
        if not product_id:
            continue
        products[product_id].append(
            {
                "product_id": product_id,
                "name": clean_text(row[1] if len(row) > 1 else ""),
                "variant_id": clean_text(row[2] if len(row) > 2 else ""),
                "variant_name": clean_text(row[3] if len(row) > 3 else ""),
                "parent_sku": clean_text(row[4] if len(row) > 4 else ""),
                "sku": clean_text(row[5] if len(row) > 5 else ""),
                "price": to_number(row[6] if len(row) > 6 else None) or 0,
                "stock": to_number(row[9] if len(row) > 9 else None),
            }
        )
    return products


def read_image_rows():
    wb = load_workbook(IMAGE_FILE, read_only=True, data_only=True)
    ws = wb.active
    images = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        product_id = clean_text(row[0] if len(row) > 0 else "")
        if not product_id:
            continue
        product_images = []
        for value in row[4:13]:
            text = clean_text(value)
            if text and text not in product_images:
                product_images.append(text)
        variant_images = {}
        for index in range(16, len(row), 2):
            name = clean_text(row[index] if index < len(row) else "")
            image = clean_text(row[index + 1] if index + 1 < len(row) else "")
            if name and image:
                variant_images[name] = image
        images[product_id] = {
            "category": clean_text(row[3] if len(row) > 3 else ""),
            "images": product_images[:9],
            "variant_images": variant_images,
        }
    return images


def read_description_rows():
    if not DESCRIPTION_FILE.exists():
        return {}
    wb = load_workbook(DESCRIPTION_FILE, read_only=True, data_only=True)
    ws = wb.active
    descriptions = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        product_id = clean_text(row[0] if len(row) > 0 else "")
        if not product_id:
            continue
        descriptions[product_id] = {
            "name": clean_text(row[2] if len(row) > 2 else ""),
            "description": clean_multiline(row[3] if len(row) > 3 else ""),
        }
    return descriptions


def read_weight_rows():
    if not WEIGHT_FILE.exists():
        return {}
    wb = load_workbook(WEIGHT_FILE, read_only=True, data_only=True)
    ws = wb.active
    weights = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        product_id = clean_text(row[0] if len(row) > 0 else "")
        variant_id = clean_text(row[3] if len(row) > 3 else "")
        if not product_id:
            continue
        weight = to_number(row[5] if len(row) > 5 else None)
        if weight is None:
            continue
        weights[(product_id, variant_id)] = weight
    return weights


def display_category(category_id: str) -> str:
    labels = {
        "nen-thom": "Nến thơm / nến tealight",
        "nen-ly": "Nến ly",
        "nen-tru": "Nến trụ",
        "combo": "Combo xông nhà",
        "bep-xong": "Bếp xông / đèn xông",
        "phu-kien": "Phụ kiện xông",
        "thao-moc-xong": "Thảo mộc xông",
        "tinh-dau": "Tinh dầu",
        "nu-tram": "Nụ trầm",
    }
    return labels.get(category_id, "Sản phẩm Phương Lâm")


def make_short_desc(product: dict, raw_description: str) -> str:
    if product.get("shortDesc"):
        return product["shortDesc"]
    source = clean_multiline(raw_description)
    source = re.sub(r"^[^\wÀ-ỹ]+", "", source)
    for line in source.splitlines():
        line = clean_text(line)
        if len(line) >= 45:
            return line[:220].rstrip(" ,.-") + ("." if not line.endswith(".") else "")
    return f"{product['name']} được chọn lọc cho nhu cầu {display_category(product['categoryId']).lower()}, phù hợp sử dụng hằng ngày và làm đẹp không gian sống."


def build_structured_description(product: dict, raw_description: str) -> str:
    raw = clean_multiline(raw_description)
    variants = [variant["name"] for variant in product.get("variants", []) if variant.get("name")]
    variant_line = ", ".join(variants[:12])
    if len(variants) > 12:
        variant_line += f", và {len(variants) - 12} phân loại khác"
    info_lines = [
        f"- Tên sản phẩm: {product['name']}",
        f"- Danh mục: {display_category(product['categoryId'])}",
    ]
    if variant_line:
        info_lines.append(f"- Phân loại: {variant_line}")
    if product.get("weight"):
        info_lines.append(f"- Cân nặng vận chuyển: {product['weight']}g")
    if product.get("sku"):
        info_lines.append(f"- Mã sản phẩm: {product['sku']}")

    intro = make_short_desc(product, raw)
    detail = raw if raw else intro
    return (
        f"{product['name'].upper()}\n\n"
        "✔ Tổng quan sản phẩm:\n"
        f"{intro}\n\n"
        "✔ Mô tả chi tiết:\n"
        f"{detail}\n\n"
        "✔ Thông tin sản phẩm:\n"
        + "\n".join(info_lines)
        + "\n\n"
        "✔ Lưu ý quan trọng:\n"
        "- Vui lòng chọn đúng phân loại trước khi đặt hàng.\n"
        "- Bảo quản sản phẩm nơi khô ráo, thoáng mát, tránh ẩm và nắng gắt.\n"
        "- Màu sắc, kích thước hoặc hình dáng có thể chênh nhẹ do ánh sáng chụp ảnh hoặc đặc tính thủ công/tự nhiên của từng sản phẩm."
    )


def build_usage(product: dict, raw_description: str) -> str:
    category_id = product.get("categoryId", "")
    name = product.get("name", "sản phẩm")
    if category_id in {"thao-moc-xong", "combo"}:
        return (
            f"HƯỚNG DẪN SỬ DỤNG {name.upper()}\n\n"
            "✔ Cách dùng:\nCho một lượng sản phẩm vừa đủ vào bếp xông hoặc dụng cụ phù hợp, làm nóng bằng nến tealight để hương thơm tỏa ra tự nhiên.\n\n"
            "✔ Liều lượng:\nDùng lượng nhỏ trước, sau đó tăng dần tùy diện tích phòng và độ đậm mùi mong muốn.\n\n"
            "✔ Sau khi dùng:\nChờ dụng cụ nguội hoàn toàn rồi vệ sinh phần thảo mộc đã xông.\n\n"
            "✔ An toàn:\nLuôn đặt bếp/dụng cụ xông trên bề mặt phẳng, chịu nhiệt và tránh xa vật dễ cháy."
        )
    if category_id in {"nen-thom", "nen-ly", "nen-tru"}:
        return (
            f"HƯỚNG DẪN SỬ DỤNG {name.upper()}\n\n"
            "✔ Cách đốt:\nĐặt nến trên bề mặt phẳng, chịu nhiệt rồi châm lửa ở tim nến.\n\n"
            "✔ Khi sử dụng:\nKhông đặt nến gần rèm, giấy, gỗ mỏng hoặc vật dễ cháy. Không di chuyển nến khi đang cháy.\n\n"
            "✔ Sau khi dùng:\nTắt nến trước khi rời khỏi phòng và chờ nến nguội hoàn toàn trước khi cất giữ.\n\n"
            "✔ Bảo quản:\nĐể nến nơi khô ráo, thoáng mát, tránh nắng nóng trực tiếp."
        )
    if category_id == "tinh-dau":
        return (
            f"HƯỚNG DẪN SỬ DỤNG {name.upper()}\n\n"
            "✔ Với đèn xông:\nNhỏ vài giọt tinh dầu vào nước trong khay đèn xông, sau đó làm nóng bằng nến hoặc điện tùy loại đèn.\n\n"
            "✔ Với máy khuếch tán:\nThêm tinh dầu theo dung tích máy và hướng dẫn của thiết bị.\n\n"
            "✔ Lưu ý:\nKhông uống tinh dầu, tránh tiếp xúc trực tiếp với mắt và để xa tầm tay trẻ em.\n\n"
            "✔ Bảo quản:\nĐậy kín nắp sau khi dùng, để nơi mát và tránh ánh nắng trực tiếp."
        )
    if category_id in {"bep-xong", "phu-kien"}:
        return (
            f"HƯỚNG DẪN SỬ DỤNG {name.upper()}\n\n"
            "✔ Cách dùng:\nĐặt sản phẩm trên bề mặt phẳng, khô ráo và lắp cùng bếp xông, nến hoặc phụ kiện phù hợp theo đúng công năng.\n\n"
            "✔ Khi sử dụng:\nNếu dùng với nhiệt hoặc nến, luôn để xa vật dễ cháy và không chạm tay trực tiếp khi sản phẩm còn nóng.\n\n"
            "✔ Vệ sinh:\nChờ sản phẩm nguội hoàn toàn rồi lau bằng khăn khô hoặc khăn ẩm nhẹ.\n\n"
            "✔ Bảo quản:\nĐể nơi khô thoáng, tránh rơi vỡ hoặc va đập mạnh."
        )
    if category_id == "nu-tram":
        return (
            f"HƯỚNG DẪN SỬ DỤNG {name.upper()}\n\n"
            "✔ Cách đốt:\nĐặt nụ trầm lên đế đốt chịu nhiệt, châm lửa phần đầu nụ rồi thổi nhẹ để nụ cháy âm ỉ.\n\n"
            "✔ Không gian phù hợp:\nDùng trong phòng khách, phòng thiền, phòng làm việc hoặc góc thư giãn.\n\n"
            "✔ An toàn:\nKhông đặt gần vật dễ cháy và không để nụ trầm cháy khi không có người trông coi.\n\n"
            "✔ Bảo quản:\nĐậy kín hộp/túi sau khi dùng, tránh ẩm để giữ hương tốt hơn."
        )
    return (
        f"HƯỚNG DẪN SỬ DỤNG {name.upper()}\n\n"
        "✔ Cách dùng:\nChọn đúng phân loại và sử dụng theo nhu cầu thực tế của gia đình.\n\n"
        "✔ Bảo quản:\nĐể nơi khô ráo, thoáng mát, tránh ẩm và ánh nắng trực tiếp.\n\n"
        "✔ Lưu ý:\nĐọc kỹ mô tả sản phẩm trước khi dùng và liên hệ shop nếu cần tư vấn thêm."
    )


def build_products():
    info = read_info_rows()
    media = read_image_rows()
    descriptions = read_description_rows()
    shipping_weights = read_weight_rows()
    products = []
    for product_id, rows in info.items():
        first = rows[0]
        media_row = media.get(product_id, {})
        name = first["name"]
        variants = []
        for row in rows:
            variant_name = row["variant_name"] or "Mặc định"
            shipping_weight = shipping_weights.get((product_id, row["variant_id"]))
            variants.append(
                {
                    "id": f"shopee_variant_{row['variant_id'] or product_id}",
                    "name": variant_name,
                    "sku": row["sku"],
                    "price": row["price"],
                    "originalPrice": None,
                    "weight": shipping_weight if shipping_weight is not None else parse_grams(variant_name),
                    "stock": row["stock"],
                    "image": media_row.get("variant_images", {}).get(variant_name, ""),
                }
            )
        prices = [variant["price"] for variant in variants if variant["price"]]
        weights = [variant["weight"] for variant in variants if variant["weight"]]
        has_real_variants = len(variants) > 1 or (variants and variants[0]["name"] != "Mặc định")
        product = {
                "id": f"shopee_{product_id}",
                "sku": first["parent_sku"] or f"SP-{product_id}",
                "categoryId": infer_category(name, media_row.get("category", "")),
                "name": name,
                "price": min(prices) if prices else 0,
                "originalPrice": None,
                "shortDesc": "",
                "description": "",
                "usage": "",
                "tag": "",
                "weight": min(weights) if weights else None,
                "images": media_row.get("images", [])[:9],
                "reviews": [],
                "variants": variants if has_real_variants else [],
        }
        description_row = descriptions.get(product_id, {})
        raw_description = description_row.get("description", "")
        if raw_description:
            product["shortDesc"] = make_short_desc(product, raw_description)
            product["description"] = build_structured_description(product, raw_description)
            product["usage"] = build_usage(product, raw_description)
        if product_id in SEO_OVERRIDES:
            override = SEO_OVERRIDES[product_id]
            for variant in product["variants"]:
                variant["name"] = override.get("variantNameMap", {}).get(variant["name"], variant["name"])
            product.update({key: value for key, value in override.items() if key != "variantNameMap"})
        products.append(product)
    return products


def main():
    products = build_products()
    OUTPUT_FILE.write_text(
        json.dumps(products, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    total_variants = sum(len(item["variants"]) for item in products)
    print(json.dumps({
        "output": str(OUTPUT_FILE),
        "products": len(products),
        "variants": total_variants,
        "sample": products[0] if products else None,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
